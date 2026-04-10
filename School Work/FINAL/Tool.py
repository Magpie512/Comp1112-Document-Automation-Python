"""
Mars Briggs
200561234
2026-03-30 1:15 PM
Game Price Point comparison
"""
import urllib.request
from bs4 import BeautifulSoup
import openpyxl
import os
import datetime
import time
import re

"""
ONE PURPOSE PER FUNCTION!                                                                                                                                                                                                                                          

Thank you Michael. Remined me of us learning try except for approx 20 seconds.
And Dom for OS confirmation
Finally, thank you to my whiteboard and to the peer tutorVictor. You may not know python but you sure know how to think good :D
Also, thanks Gillian for helping with list managment

"""

# Globals
priceBook = None
priceSheet = None
websiteList = []
gameTitleList = [] #in hindsight i should have just named it gameList

# User Input ===================================================================================

# User guide and warning
def disclose():
    print("This program compares game prices between websites.")
    print("Provide sites that show game titles AND prices.")
    print("Sites requiring login will not work.\n")

# Asks user for sites to use inside of a while loop to ensure multiple sites if needed
def siteInput():
    addingMore = True
    while addingMore:
        site = input("Enter a website: ").strip()
        websiteList.append(site)
        print("Added " + site + " to the list.")
        print("Would you like to add another? (Y/N)")
        if input().strip().upper() == "N":
            addingMore = False

# Asks user for games to use inside of the same logic as above
def gameInput():
    addingMore = True
    while addingMore:
        game = input("Enter a game title: ").strip()
        gameTitleList.append(game)
        print("Added " + game + " to the list.")
        print("Would you like to add another? (Y/N)")
        if input().strip().upper() == "N":                                                                                                                                                                                                                                                                                                                                   
            addingMore = False

# Excel Setup Functions ===========================================================================================

# Initializes the Excel workbook
def createWorkbook():
    global priceBook, priceSheet
    priceBook = openpyxl.Workbook()
    priceSheet = priceBook.active

# Creates the column headers
def excelHeader():
    priceSheet["A1"] = "Game(s)"
    priceSheet["B1"] = "Site(s)"
    priceSheet["C1"] = "Price(s)"

# Scraping Logic Functions ==========================================================================================

# Scrape the price
def scrapePrice(targetSite, targetGame):
    pageText = fetchPage(targetSite)
    if pageText is None:
        return "Error reaching site"
    if not findGame(pageText, targetGame):
        return "Game not found"
    return findPrice(pageText)

# Fetches the page
def fetchPage(targetSite):
    try:
        html = urllib.request.urlopen(targetSite).read()
        soup = BeautifulSoup(html, "html.parser")
        return soup.get_text()
    except Exception as e:
        print("Error reaching " + targetSite + ": " + str(e))
        return None

# Finds the specific game
def findGame(pageText, targetGame):
    return targetGame.lower() in pageText.lower()

# Does it look like a pricetag function
# def isPriceSymbol(word):
#     # My weakass dollar finder logic which i now realise locks me into an american market but thats all my gf buys from so i t hink its fine
#     if len(word) > 1:
#         if word[0] == "$":
#             return True
#     return False

# Locates the price using the function above to find the symbol and thus the price
# def findPrice(pageText):
#     words = pageText.split()
#     for word in words:
#         if isPriceSymbol(word):
#             return 
#     return "No price found"

# Regex that "improves" the above function. it currently finds first on page
import re

def findPrice(pageText):
    # Regex for price patterns
    pricePattern = r"\$\d{2}\.\d{2}"

    match = re.search(pricePattern, pageText)
    if match:
        return match[0].strip()

    return "No price found"

# Excel Write =====================================================================================================

# Populates the rows of the newly created Excel sheet
def writeToExcel(resultsList):
    rowIndex = 2
    for game, site, price in resultsList:
        priceSheet.cell(row=rowIndex, column=1).value = game
        priceSheet.cell(row=rowIndex, column=2).value = site
        priceSheet.cell(row=rowIndex, column=3).value = price
        rowIndex += 1 # christ did we learn this in class? can I assume this?

# Save le file to excel
def saveExcel():
    try:
        priceBook.save("gamePrices.xlsx")

    except PermissionError as e:
        print("Error: " + str(e))

# deletes the existing file file its there
def deleteExistingFile():
    if os.path.exists("gamePrices.xlsx"):
        try:
            os.remove("gamePrices.xlsx")

        except PermissionError as e:
            print("Error: " + str(e))

# Utility Functions =======================================================================================================

# Notifies the user of a successful save
def notifyUser():
    print("Results saved to gamePrices.xlsx")

# Opens the saved file via the default program ( you really should have one assigned anyway)
def openProgram():
    try:
        os.startfile("gamePrices.xlsx")
    except Exception:
        print("Could not open file automatically.\nJust set a default program, man.")

# Delays the execution again until the next day
def waitUntilMidnight():
    now = datetime.datetime.now()
    tomorrow = datetime.datetime(now.year, now.month, now.day) + datetime.timedelta(days=1)
    # Calculate exact sleep duration
    secondsToWait = (tomorrow - now).total_seconds()
    time.sleep(secondsToWait)

# Main Logic und Execution ===============================================================================================

disclose()
input("Press Enter to begin setup")
siteInput()
gameInput()

# Initial Setup
createWorkbook()
excelHeader()

# Sets when the loop will end
endDate = datetime.datetime.now() + datetime.timedelta(days=7)

# Main Loop
while datetime.datetime.now() < endDate:
    cycleResults = []

    for gameTitle in gameTitleList:
        for website in websiteList:
            foundPrice = scrapePrice(website, gameTitle)
            cycleResults.append([gameTitle, website, foundPrice]) # thank you Gillian for showing me the way during one of our classes

    writeToExcel(cycleResults)
    saveExcel()
    notifyUser()
    openProgram()
    
    # Wait until the start of the next day, then clear and recreate the workbook
    waitUntilMidnight()
    deleteExistingFile()
    createWorkbook()
    excelHeader()